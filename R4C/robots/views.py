from datetime import datetime, timedelta
from json import JSONDecodeError, loads
from openpyxl import Workbook
from django.views.decorators.csrf import csrf_exempt

from django.http import (
    HttpResponse,
    HttpResponseBadRequest,
    FileResponse,
)

from .orm_functions import create_robot, get_unique_version_dicts
from .validators import validate_json, validate_datetime_field


@csrf_exempt
def create_robots_api(request):
    """Функциональное вью для API-endpoint'а, принимающий JSON-объект POST-методом
    с информацией о созданных роботах с последующей их валидацией и сохранением в базу данных."""

    if request.method == 'POST':
        try:
            data = loads(request.body)
        # если прилетел кривой JSON
        except JSONDecodeError as er:
            return HttpResponseBadRequest(
                f'Cannot decode JSON: {er}',
                status=400,
            )

        # если в JSON'е один робот
        if type(data) == dict:
            if validate_json(json_data=data) and validate_datetime_field(datetime_text=data['created']):
                create_robot(info=data)
                return HttpResponse(
                    'New object was successfully added to database.',
                    status=201,
                )
            else:
                return HttpResponseBadRequest(
                    'JSON object is invalid.',
                    status=400,
                )

        # если >1 робота
        if type(data) == list:
            valid_objs = 0
            invalid_objs = 0

            for item in data:
                if validate_json(json_data=item) and validate_datetime_field(datetime_text=item['created']):
                    create_robot(info=item)
                    valid_objs += 1
                else:
                    # просто выведем на странице кол-во невалидных объектов
                    # вдруг какие-то другие нормальные и смогут сохраниться в БД
                    invalid_objs += 1

            return HttpResponse(
                f'{valid_objs} valid objects added to database.\n'
                f'{invalid_objs} invalid objects ignored.',
                status=201,
            )

    else:
        return HttpResponseBadRequest(
            'You need to pass your JSON data via POST-method.',
            status=405,
        )


@csrf_exempt
def download_weekly_excel(request):
    """Функциональное вью для URL-адреса, позволяющего скачать Excel-файл со статистикой
    произведённых за последнюю неделю роботов."""

    if request.method == 'GET':
        today = datetime.now()
        week_ago = today - timedelta(weeks=1)
        query_dicts = get_unique_version_dicts(time_filter=week_ago)

        # если ни одного робота за последнюю неделю
        if not query_dicts:
            return HttpResponse(
                'No robots were created in last week.',
                status=200,
            )

        file_name = f"""results/{week_ago.strftime("%Y-%m-%d %H-%M")}-{today.strftime("%Y-%m-%d %H-%M")}.xlsx"""
        xl = Workbook()

        for iter_dict in query_dicts:
            model_name = iter_dict['model']

            # если нет листа с названием модели, то создаём
            if model_name not in xl.sheetnames:
                new_sheet = xl.create_sheet(title=model_name)
                columns = [
                    'Модель',
                    'Версия',
                    'Количество за неделю',
                ]
                new_sheet.append(columns)

            # в нужном листе добавляем значения из итерируемого словаря
            cur_sheet = xl[model_name]
            cur_sheet.append([val for val in iter_dict.values()])

        # удаляем стандартный первый лист Sheet1 - он не нужен
        xl.remove(worksheet=xl[xl.sheetnames[0]])
        xl.save(filename=file_name)

        return FileResponse(
            open(file=file_name, mode='rb'),
            status=200,
        )

    else:
        return HttpResponseBadRequest(
            'You need to access this page via GET-method.',
            status=405,
        )
